"""
Endpoint dasar akuntansi: chart of accounts (akun) dan jurnal umum.

Catatan: endpoint di sini baru mencakup pencatatan (input) data.
Perhitungan buku besar, neraca saldo, jurnal penyesuaian, dan laporan
keuangan (laba rugi, posisi keuangan, CALK) akan ditambahkan di
app/accounting/*.py pada tahap "accounting engine" berikutnya, lalu
di-expose lewat router ini juga.
"""

import csv
import io
from datetime import date

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.accounting import calk as calk_engine
from app.accounting import dashboard_metrics
from app.accounting import jurnal_penyesuaian
from app.accounting import laporan_laba_rugi as laba_rugi_engine
from app.accounting import laporan_posisi_keuangan as posisi_keuangan_engine
from app.accounting import neraca_saldo as neraca_saldo_engine
from app.accounting import tax_engine
from app.accounting.buku_besar import get_buku_besar
from app.accounting.jurnal_umum import JurnalDetailInputDTO, JurnalError
from app.accounting.jurnal_umum import buat_jurnal as service_buat_jurnal
from app.accounting.jurnal_umum import list_jurnal as service_list_jurnal
from app.config.logging import get_logger
from app.database.database import get_db
from app.database.models import Akun, JenisJurnal, JurnalUmum, User
from app.middleware.auth import create_download_token, require_active_user
from app.schemas.report_schema import (
    BukuBesarResponse,
    CalkResponse,
    LaporanLabaRugiResponse,
    LaporanPosisiKeuanganResponse,
    NeracaSaldoResponse,
    PenyesuaianPenyusutanInput,
    PenyesuaianPerlengkapanInput,
    PPhFinalUMKMRequest,
    PPhFinalUMKMResponse,
    PPNRequest,
    PPNResponse,
)
from app.schemas.transaction_schema import (
    AkunCreate,
    AkunResponse,
    JurnalDetailResponse,
    JurnalUmumCreate,
    JurnalUmumResponse,
)

router = APIRouter(prefix="/accounting", tags=["Accounting"])
logger = get_logger(__name__)


# ---------------------------------------------------------------------------
# Chart of Accounts
# ---------------------------------------------------------------------------
@router.get("/akun", response_model=list[AkunResponse])
def list_akun(
    kategori: str | None = None,
    db: Session = Depends(get_db),
    _: User = Depends(require_active_user),
) -> list[Akun]:
    query = db.query(Akun).filter(Akun.is_active.is_(True))
    if kategori:
        query = query.filter(Akun.kategori == kategori.upper())
    return query.order_by(Akun.kode_akun).all()


@router.post("/akun", response_model=AkunResponse, status_code=status.HTTP_201_CREATED)
def create_akun(
    payload: AkunCreate,
    db: Session = Depends(get_db),
    _: User = Depends(require_active_user),
) -> Akun:
    existing = db.query(Akun).filter(Akun.kode_akun == payload.kode_akun).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"Kode akun {payload.kode_akun} sudah dipakai",
        )
    akun = Akun(**payload.model_dump())
    db.add(akun)
    db.commit()
    db.refresh(akun)
    return akun


# ---------------------------------------------------------------------------
# Jurnal Umum
# ---------------------------------------------------------------------------
def _to_jurnal_response(jurnal: JurnalUmum) -> JurnalUmumResponse:
    detail_responses = [
        JurnalDetailResponse(
            kode_akun=d.akun.kode_akun,
            nama_akun=d.akun.nama_akun,
            debit=float(d.debit),
            kredit=float(d.kredit),
            keterangan=d.keterangan,
        )
        for d in jurnal.detail
    ]
    return JurnalUmumResponse(
        id=jurnal.id,
        no_bukti=jurnal.no_bukti,
        tanggal=jurnal.tanggal,
        deskripsi=jurnal.deskripsi,
        detail=detail_responses,
        created_at=jurnal.created_at,
    )


@router.get("/jurnal", response_model=list[JurnalUmumResponse])
def list_jurnal(
    tanggal_mulai: date | None = None,
    tanggal_akhir: date | None = None,
    jenis: JenisJurnal | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_active_user),
) -> list[JurnalUmumResponse]:
    jurnal_list = service_list_jurnal(db, tanggal_mulai, tanggal_akhir, jenis, user_id=current_user.id)
    return [_to_jurnal_response(j) for j in jurnal_list]


@router.post("/jurnal", response_model=JurnalUmumResponse, status_code=status.HTTP_201_CREATED)
def create_jurnal(
    payload: JurnalUmumCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_active_user),
) -> JurnalUmumResponse:
    detail_dto = [
        JurnalDetailInputDTO(d.kode_akun, d.debit, d.kredit, d.keterangan) for d in payload.detail
    ]
    try:
        jurnal = service_buat_jurnal(
            db,
            no_bukti=payload.no_bukti,
            tanggal=payload.tanggal,
            deskripsi=payload.deskripsi,
            detail=detail_dto,
            created_by_id=current_user.id,
        )
    except JurnalError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
    return _to_jurnal_response(jurnal)


# ---------------------------------------------------------------------------
# Jurnal Penyesuaian
# ---------------------------------------------------------------------------
@router.post("/jurnal-penyesuaian/perlengkapan", response_model=JurnalUmumResponse, status_code=status.HTTP_201_CREATED)
def create_penyesuaian_perlengkapan(
    payload: PenyesuaianPerlengkapanInput,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_active_user),
) -> JurnalUmumResponse:
    try:
        jurnal = jurnal_penyesuaian.penyesuaian_perlengkapan_terpakai(
            db,
            no_bukti=payload.no_bukti,
            tanggal=payload.tanggal,
            nilai_terpakai=payload.nilai_terpakai,
            created_by_id=current_user.id,
            kode_akun_perlengkapan=payload.kode_akun_perlengkapan,
            kode_akun_beban_perlengkapan=payload.kode_akun_beban_perlengkapan,
        )
    except JurnalError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
    return _to_jurnal_response(jurnal)


@router.post("/jurnal-penyesuaian/penyusutan", response_model=JurnalUmumResponse, status_code=status.HTTP_201_CREATED)
def create_penyesuaian_penyusutan(
    payload: PenyesuaianPenyusutanInput,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_active_user),
) -> JurnalUmumResponse:
    try:
        jurnal = jurnal_penyesuaian.penyesuaian_penyusutan_aset(
            db,
            no_bukti=payload.no_bukti,
            tanggal=payload.tanggal,
            harga_perolehan=payload.harga_perolehan,
            nilai_residu=payload.nilai_residu,
            umur_ekonomis_tahun=payload.umur_ekonomis_tahun,
            created_by_id=current_user.id,
            kode_akun_beban_penyusutan=payload.kode_akun_beban_penyusutan,
            kode_akun_akumulasi_penyusutan=payload.kode_akun_akumulasi_penyusutan,
            bulan_penyusutan=payload.bulan_penyusutan,
        )
    except JurnalError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
    return _to_jurnal_response(jurnal)


# ---------------------------------------------------------------------------
# Buku Besar
# ---------------------------------------------------------------------------
@router.get("/buku-besar/{kode_akun}", response_model=BukuBesarResponse)
def buku_besar(
    kode_akun: str,
    tanggal_mulai: date | None = None,
    tanggal_akhir: date | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_active_user),
) -> BukuBesarResponse:
    try:
        hasil = get_buku_besar(db, kode_akun, tanggal_mulai, tanggal_akhir, user_id=current_user.id)
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(exc)) from exc
    return BukuBesarResponse(**hasil.__dict__)


# ---------------------------------------------------------------------------
# Neraca Saldo
# ---------------------------------------------------------------------------
@router.get("/laporan/neraca-saldo", response_model=NeracaSaldoResponse)
def neraca_saldo(
    tanggal_per: date | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_active_user),
) -> NeracaSaldoResponse:
    hasil = neraca_saldo_engine.get_neraca_saldo(db, tanggal_per, user_id=current_user.id)
    return NeracaSaldoResponse(
        tanggal_per=hasil.tanggal_per,
        baris=hasil.baris,
        total_debit=hasil.total_debit,
        total_kredit=hasil.total_kredit,
        is_balance=hasil.is_balance,
    )


# ---------------------------------------------------------------------------
# Laporan Laba Rugi
# ---------------------------------------------------------------------------
@router.get("/laporan/laba-rugi", response_model=LaporanLabaRugiResponse)
def laporan_laba_rugi(
    tanggal_per: date | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_active_user),
) -> LaporanLabaRugiResponse:
    try:
        hasil = laba_rugi_engine.get_laporan_laba_rugi(db, tanggal_per, user_id=current_user.id)
    except Exception as exc:
        logger.error("Gagal menghitung laporan laba rugi: %s", exc)
        raise HTTPException(status_code=500, detail=f"Gagal menghitung laporan laba rugi: {exc}") from exc
    return LaporanLabaRugiResponse(
        tanggal_per=hasil.tanggal_per,
        pendapatan=hasil.pendapatan,
        hpp=hasil.hpp,
        beban_operasional=hasil.beban_operasional,
        total_pendapatan=hasil.total_pendapatan,
        total_hpp=hasil.total_hpp,
        laba_kotor=hasil.laba_kotor,
        total_beban_operasional=hasil.total_beban_operasional,
        laba_bersih=hasil.laba_bersih,
    )


# ---------------------------------------------------------------------------
# Laporan Posisi Keuangan
# ---------------------------------------------------------------------------
@router.get("/laporan/posisi-keuangan", response_model=LaporanPosisiKeuanganResponse)
def laporan_posisi_keuangan(
    tanggal_per: date | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_active_user),
) -> LaporanPosisiKeuanganResponse:
    try:
        hasil = posisi_keuangan_engine.get_laporan_posisi_keuangan(db, tanggal_per, user_id=current_user.id)
    except Exception as exc:
        logger.error("Gagal menghitung laporan posisi keuangan: %s", exc)
        raise HTTPException(status_code=500, detail=f"Gagal menghitung laporan posisi keuangan: {exc}") from exc
    return LaporanPosisiKeuanganResponse(
        tanggal_per=hasil.tanggal_per,
        aset=hasil.aset,
        liabilitas=hasil.liabilitas,
        modal=hasil.modal,
        laba_rugi_berjalan=hasil.laba_rugi_berjalan,
        total_aset=hasil.total_aset,
        total_liabilitas=hasil.total_liabilitas,
        total_modal=hasil.total_modal,
        total_liabilitas_dan_modal=hasil.total_liabilitas_dan_modal,
        is_balance=hasil.is_balance,
    )


# ---------------------------------------------------------------------------
# CALK
# ---------------------------------------------------------------------------
@router.get("/laporan/calk", response_model=CalkResponse)
def laporan_calk(
    tanggal_per: date | None = None,
    ambang_material: float = 0,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_active_user),
) -> CalkResponse:
    try:
        hasil = calk_engine.get_calk(db, tanggal_per, ambang_material, user_id=current_user.id)
    except Exception as exc:
        logger.error("Gagal menghitung CALK: %s", exc)
        raise HTTPException(status_code=500, detail=f"Gagal menghitung CALK: {exc}") from exc
    return CalkResponse(
        tanggal_per=hasil.tanggal_per,
        kebijakan_akuntansi=hasil.kebijakan_akuntansi,
        rincian_aset=hasil.rincian_aset,
        rincian_liabilitas=hasil.rincian_liabilitas,
        rincian_pendapatan=hasil.rincian_pendapatan,
        rincian_beban=hasil.rincian_beban,
        catatan_tambahan=hasil.catatan_tambahan,
    )


# ---------------------------------------------------------------------------
# Pajak (PPh Final UMKM & PPN)
# ---------------------------------------------------------------------------
@router.post("/pajak/pph-final-umkm", response_model=PPhFinalUMKMResponse)
def hitung_pph_final_umkm(
    payload: PPhFinalUMKMRequest,
    _: User = Depends(require_active_user),
) -> PPhFinalUMKMResponse:
    hasil = tax_engine.hitung_pph_final_umkm(
        omzet_bulan_ini=payload.omzet_bulan_ini,
        omzet_kumulatif_sebelum_bulan_ini=payload.omzet_kumulatif_sebelum_bulan_ini,
        wp_orang_pribadi=payload.wp_orang_pribadi,
    )
    return PPhFinalUMKMResponse(**hasil.__dict__)


@router.post("/pajak/ppn", response_model=PPNResponse)
def hitung_ppn(
    payload: PPNRequest,
    _: User = Depends(require_active_user),
) -> PPNResponse:
    if payload.sudah_termasuk_ppn:
        hasil = tax_engine.hitung_ppn_dari_harga_termasuk_pajak(payload.nilai, payload.barang_mewah)
    else:
        hasil = tax_engine.hitung_ppn(payload.nilai, payload.barang_mewah)
    return PPNResponse(**hasil.__dict__)


# ---------------------------------------------------------------------------
# Export Laporan (Excel / CSV)
# ---------------------------------------------------------------------------
REPORT_TITLES = {
    "neraca-saldo": "Laporan Neraca Saldo",
    "laba-rugi": "Laporan Laba Rugi",
    "posisi-keuangan": "Laporan Posisi Keuangan",
    "calk": "Catatan Atas Laporan Keuangan (CALK)",
}


def _format_date(date_str: date | None) -> str:
    if not date_str:
        return "-"
    return date_str.strftime("%d %B %Y")


def _build_neraca_saldo_rows(data) -> list[list]:
    header = ["Kode Akun", "Nama Akun", "Kategori", "Debit", "Kredit"]
    rows = [header]
    for r in data.baris:
        rows.append([r.kode_akun, r.nama_akun, r.kategori, r.debit, r.kredit])
    rows.append(["", "", "TOTAL", data.total_debit, data.total_kredit])
    return rows


def _build_laba_rugi_rows(data) -> list[list]:
    header = ["Kode Akun", "Nama Akun", "Nilai"]
    rows = [header, ["", "PENDAPATAN", ""]]
    for r in data.pendapatan:
        rows.append([r.kode_akun, r.nama_akun, r.nilai])
    rows.append(["", "Total Pendapatan", data.total_pendapatan])
    rows.append(["", "", ""])
    rows.append(["", "HARGA POKOK PENJUALAN (HPP)", ""])
    for r in data.hpp:
        rows.append([r.kode_akun, r.nama_akun, r.nilai])
    rows.append(["", "Total HPP", data.total_hpp])
    rows.append(["", "LABA KOTOR", data.laba_kotor])
    rows.append(["", "", ""])
    rows.append(["", "BEBAN OPERASIONAL", ""])
    for r in data.beban_operasional:
        rows.append([r.kode_akun, r.nama_akun, r.nilai])
    rows.append(["", "Total Beban Operasional", data.total_beban_operasional])
    rows.append(["", "", ""])
    rows.append(["", "LABA BERSIH", data.laba_bersih])
    return rows


def _build_posisi_keuangan_rows(data) -> list[list]:
    header = ["Kode Akun", "Nama Akun", "Nilai"]
    rows = [header, ["", "ASET", ""]]
    for r in data.aset:
        rows.append([r.kode_akun, r.nama_akun, r.nilai])
    rows.append(["", "Total Aset", data.total_aset])
    rows.append(["", "", ""])
    rows.append(["", "LIABILITAS", ""])
    for r in data.liabilitas:
        rows.append([r.kode_akun, r.nama_akun, r.nilai])
    rows.append(["", "Total Liabilitas", data.total_liabilitas])
    rows.append(["", "", ""])
    rows.append(["", "MODAL", ""])
    for r in data.modal:
        rows.append([r.kode_akun, r.nama_akun, r.nilai])
    rows.append(["", "Total Modal", data.total_modal])
    rows.append(["", "Laba/Rugi Berjalan", data.laba_rugi_berjalan])
    rows.append(["", "TOTAL LIABILITAS DAN MODAL", data.total_liabilitas_dan_modal])
    return rows


def _build_calk_rows(data) -> list[list]:
    header = ["Kode Akun", "Nama Akun", "Catatan", "Nilai"]
    rows = [header]
    if data.kebijakan_akuntansi:
        rows.append(["", "KEBIJAKAN AKUNTANSI", "", ""])
        for k in data.kebijakan_akuntansi:
            rows.append(["", k, "", ""])
        rows.append(["", "", "", ""])
    for label, items in [
        ("RINCIAN ASET", data.rincian_aset),
        ("RINCIAN LIABILITAS", data.rincian_liabilitas),
        ("RINCIAN PENDAPATAN", data.rincian_pendapatan),
        ("RINCIAN BEBAN", data.rincian_beban),
    ]:
        if items:
            rows.append(["", label, "", ""])
            for r in items:
                rows.append([r.kode_akun, r.nama_akun, r.catatan or "", r.nilai])
            rows.append(["", "", "", ""])
    if data.catatan_tambahan:
        rows.append(["", "CATATAN TAMBAHAN", "", ""])
        for c in data.catatan_tambahan:
            rows.append(["", c, "", ""])
    return rows


def _build_rows(report_type: str, data) -> list[list]:
    builders = {
        "neraca-saldo": _build_neraca_saldo_rows,
        "laba-rugi": _build_laba_rugi_rows,
        "posisi-keuangan": _build_posisi_keuangan_rows,
        "calk": _build_calk_rows,
    }
    return builders[report_type](data)


def _generate_xlsx(report_type: str, data, tanggal_per: date | None) -> io.BytesIO:
    rows = _build_rows(report_type, data)
    wb = Workbook()
    ws = wb.active
    ws.title = REPORT_TITLES.get(report_type, "Laporan")

    purple_fill = PatternFill(start_color="4B2B85", end_color="4B2B85", fill_type="solid")
    purple_light = PatternFill(start_color="F5F3FF", end_color="F5F3FF", fill_type="solid")
    white_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    bold_font = Font(name="Calibri", bold=True, size=11)
    normal_font = Font(name="Calibri", size=10)
    section_font = Font(name="Calibri", bold=True, size=11, color="4B2B85")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    title_cell = ws.cell(row=1, column=1, value=REPORT_TITLES.get(report_type, "Laporan"))
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="4B2B85")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    ws.cell(row=2, column=1, value=f"Per {_format_date(tanggal_per)}")

    start_row = 4
    for row_idx, row_data in enumerate(rows):
        excel_row = start_row + row_idx
        is_header = row_idx == 0
        is_section = not is_header and len(row_data) >= 2 and row_data[1] and not row_data[0] and any(
            keyword in str(row_data[1]).upper()
            for keyword in ["PENDAPATAN", "HPP", "BEBAN", "ASET", "LIABILITAS", "MODAL", "KEBIJAKAN", "RINCIAN", "CATATAN", "TOTAL"]
        )
        is_total = not is_header and any(
            str(row_data[i]).upper().startswith("TOTAL") or str(row_data[i]).upper().startswith("LABA")
            for i in range(len(row_data))
            if isinstance(row_data[i], str)
        )

        for col_idx, value in enumerate(row_data):
            cell = ws.cell(row=excel_row, column=col_idx + 1, value=value if value != "" else None)
            cell.border = thin_border
            cell.font = normal_font

            if is_header:
                cell.fill = purple_fill
                cell.font = white_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif is_section:
                cell.font = section_font
                cell.fill = purple_light
            elif is_total:
                cell.font = bold_font
                cell.fill = purple_light
            elif col_idx >= 3 and isinstance(value, (int, float)) and value != 0:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")

    for col_idx in range(1, len(rows[0]) + 1 if rows else 1):
        max_length = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=start_row, max_row=ws.max_row):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _generate_csv(report_type: str, data, tanggal_per: date | None) -> io.StringIO:
    rows = _build_rows(report_type, data)
    output = io.StringIO()
    output.write("\ufeff")  # UTF-8 BOM untuk kompatibilitas Excel
    writer = csv.writer(output)
    for row in rows:
        writer.writerow(row)
    output.seek(0)
    return output


# ---------------------------------------------------------------------------
# Export ALL Laporan (combined)
# ---------------------------------------------------------------------------
ALL_REPORTS = ["neraca-saldo", "laba-rugi", "posisi-keuangan", "calk"]


def _generate_all_xlsx(all_data: dict, tanggal_per: date | None) -> io.BytesIO:
    wb = Workbook()
    wb.remove(wb.active)

    purple_fill = PatternFill(start_color="4B2B85", end_color="4B2B85", fill_type="solid")
    purple_light = PatternFill(start_color="F5F3FF", end_color="F5F3FF", fill_type="solid")
    white_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    bold_font = Font(name="Calibri", bold=True, size=11)
    normal_font = Font(name="Calibri", size=10)
    section_font = Font(name="Calibri", bold=True, size=11, color="4B2B85")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    for report_type in ALL_REPORTS:
        data = all_data.get(report_type)
        if not data:
            continue

        title = REPORT_TITLES.get(report_type, "Laporan")
        sheet_name = title.replace("Laporan ", "").replace("Catatan Atas ", "").replace(" (CALK)", "")[:31]
        ws = wb.create_sheet(title=sheet_name)

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(name="Calibri", bold=True, size=14, color="4B2B85")

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
        ws.cell(row=2, column=1, value=f"Per {_format_date(tanggal_per)}")

        rows = _build_rows(report_type, data)
        start_row = 4
        for row_idx, row_data in enumerate(rows):
            excel_row = start_row + row_idx
            is_header = row_idx == 0
            is_section = not is_header and len(row_data) >= 2 and row_data[1] and not row_data[0] and any(
                keyword in str(row_data[1]).upper()
                for keyword in ["PENDAPATAN", "HPP", "BEBAN", "ASET", "LIABILITAS", "MODAL", "KEBIJAKAN", "RINCIAN", "CATATAN", "TOTAL"]
            )
            is_total = not is_header and any(
                isinstance(row_data[i], str) and (
                    str(row_data[i]).upper().startswith("TOTAL") or str(row_data[i]).upper().startswith("LABA")
                )
                for i in range(len(row_data))
            )

            for col_idx, value in enumerate(row_data):
                cell = ws.cell(row=excel_row, column=col_idx + 1, value=value if value != "" else None)
                cell.border = thin_border
                cell.font = normal_font

                if is_header:
                    cell.fill = purple_fill
                    cell.font = white_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif is_section:
                    cell.font = section_font
                    cell.fill = purple_light
                elif is_total:
                    cell.font = bold_font
                    cell.fill = purple_light
                elif col_idx >= 3 and isinstance(value, (int, float)) and value != 0:
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="right")

        for col_idx in range(1, len(rows[0]) + 1 if rows else 1):
            max_length = 0
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=start_row, max_row=ws.max_row):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _generate_all_csv(all_data: dict, tanggal_per: date | None) -> io.StringIO:
    output = io.StringIO()
    output.write("\ufeff")  # UTF-8 BOM untuk kompatibilitas Excel
    writer = csv.writer(output)

    for idx, report_type in enumerate(ALL_REPORTS):
        data = all_data.get(report_type)
        if not data:
            continue

        if idx > 0:
            writer.writerow([])
            writer.writerow([])

        title = REPORT_TITLES.get(report_type, "Laporan")
        writer.writerow([title])
        writer.writerow([f"Per {_format_date(tanggal_per)}"])
        writer.writerow([])

        rows = _build_rows(report_type, data)
        for row in rows:
            writer.writerow(row)

    output.seek(0)
    return output


@router.get("/laporan/{report_type}/export")
def export_laporan(
    report_type: str,
    format: str = Query("xlsx", regex="^(xlsx|csv)$"),
    tanggal_per: date | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_active_user),
):
    if report_type not in REPORT_TITLES:
        raise HTTPException(status_code=404, detail=f"Jenis laporan '{report_type}' tidak dikenal")

    engines = {
        "neraca-saldo": neraca_saldo_engine.get_neraca_saldo,
        "laba-rugi": laba_rugi_engine.get_laporan_laba_rugi,
        "posisi-keuangan": posisi_keuangan_engine.get_laporan_posisi_keuangan,
        "calk": calk_engine.get_calk,
    }

    try:
        data = engines[report_type](db, tanggal_per, user_id=current_user.id)
    except Exception as exc:
        logger.error("Gagal menghitung laporan %s untuk export: %s", report_type, exc)
        raise HTTPException(status_code=500, detail=f"Gagal menghitung laporan: {exc}") from exc

    safe_name = report_type.replace(" ", "_")
    date_str = tanggal_per.strftime("%Y%m%d") if tanggal_per else date.today().strftime("%Y%m%d")
    filename = f"{safe_name}_{date_str}"

    if format == "xlsx":
        content = _generate_xlsx(report_type, data, tanggal_per)
        return StreamingResponse(
            content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
        )
    else:
        content = _generate_csv(report_type, data, tanggal_per)
        return StreamingResponse(
            iter([content.getvalue()]),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename}.csv"'},
        )


@router.get("/laporan/export-all")
def export_all_laporan(
    format: str = Query("xlsx", regex="^(xlsx|csv)$"),
    tanggal_per: date | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_active_user),
):
    engines = {
        "neraca-saldo": neraca_saldo_engine.get_neraca_saldo,
        "laba-rugi": laba_rugi_engine.get_laporan_laba_rugi,
        "posisi-keuangan": posisi_keuangan_engine.get_laporan_posisi_keuangan,
        "calk": calk_engine.get_calk,
    }

    all_data = {}
    for rt, engine_fn in engines.items():
        try:
            all_data[rt] = engine_fn(db, tanggal_per, user_id=current_user.id)
        except Exception as exc:
            logger.error("Gagal menghitung laporan %s untuk export-all: %s", rt, exc)
            raise HTTPException(status_code=500, detail=f"Gagal menghitung laporan {rt}: {exc}") from exc

    date_str = tanggal_per.strftime("%Y%m%d") if tanggal_per else date.today().strftime("%Y%m%d")
    filename = f"laporan-keuangan-semua_{date_str}"

    if format == "xlsx":
        content = _generate_all_xlsx(all_data, tanggal_per)
        return StreamingResponse(
            content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
        )
    else:
        content = _generate_all_csv(all_data, tanggal_per)
        return StreamingResponse(
            iter([content.getvalue()]),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename}.csv"'},
        )


@router.get("/health")
def health() -> dict:
    return {"status": "ok", "module": "accounting"}
